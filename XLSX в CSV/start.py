import os
import openpyxl
import csv

# Путь к папке с исходными xlsx-файлами
FOLDER_PATH = '.'  # текущая папка
OUTPUT_FOLDER = 'csv_output'  # папка для сохранения CSV-файлов

# Создаём папку для CSV, если её нет
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Перебираем все .xlsx файлы в папке
for filename in os.listdir(FOLDER_PATH):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(FOLDER_PATH, filename)
        print(f"\n📄 Обрабатываем файл: {filename}")

        try:
            workbook = openpyxl.load_workbook(file_path)

            # Для каждого листа в Excel-файле создаём отдельный CSV-файл
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                csv_filename = f"{os.path.splitext(filename)[0]}_{sheet_name}.csv"
                csv_path = os.path.join(OUTPUT_FOLDER, csv_filename)

                print(f"💾 Сохраняю лист '{sheet_name}' как: {csv_filename}")

                with open(csv_path, mode='w', encoding='utf-8', newline='') as f:
                    writer = csv.writer(f)

                    # Перебираем строки и записываем в CSV
                    for row in sheet.iter_rows(values_only=True):
                        writer.writerow(row)

            print(f"✅ Файл {filename} успешно конвертирован в CSV.")

        except Exception as e:
            print(f"❌ Ошибка при обработке файла {filename}: {e}")

print(f"\n✨ Все файлы конвертированы. Найди их в папке: '{OUTPUT_FOLDER}'")