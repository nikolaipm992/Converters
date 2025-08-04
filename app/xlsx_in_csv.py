import os
import openpyxl
import csv

# Фиксированный путь к папке Converters
FOLDER_PATH = r'D:\Converters'

# Проверяем существование папки
if not os.path.exists(FOLDER_PATH):
    print(f"❌ Папка {FOLDER_PATH} не существует!")
    input("Нажмите Enter для выхода...")
    exit()

print(f"🔍 Поиск .xlsx файлов в папке: {FOLDER_PATH}")
xlsx_files_found = 0

# Перебираем все .xlsx файлы в папке
for filename in os.listdir(FOLDER_PATH):
    if filename.endswith('.xlsx') and not filename.startswith('~$'):  # игнорируем временные файлы
        xlsx_files_found += 1
        file_path = os.path.join(FOLDER_PATH, filename)
        print(f"\n📄 Обрабатываем файл: {filename}")

        try:
            workbook = openpyxl.load_workbook(file_path)

            # Для каждого листа в Excel-файле создаём отдельный CSV-файл
            sheets_converted = 0
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                csv_filename = f"{os.path.splitext(filename)[0]}_{sheet_name}.csv"
                csv_path = os.path.join(FOLDER_PATH, csv_filename)  # Сохраняем直接 в FOLDER_PATH

                print(f"💾 Сохраняю лист '{sheet_name}' как: {csv_filename}")

                with open(csv_path, mode='w', encoding='utf-8', newline='', errors='ignore') as f:
                    writer = csv.writer(f)

                    # Перебираем строки и записываем в CSV
                    rows_written = 0
                    for row in sheet.iter_rows(values_only=True):
                        # Фильтруем пустые строки
                        if any(cell is not None for cell in row):
                            writer.writerow(row)
                            rows_written += 1

                print(f"✅ Лист '{sheet_name}' сохранен ({rows_written} строк)")
                sheets_converted += 1

            print(f"✅ Файл {filename} успешно конвертирован в CSV ({sheets_converted} листов).")

        except Exception as e:
            print(f"❌ Ошибка при обработке файла {filename}: {e}")

if xlsx_files_found == 0:
    print("❌ В папке D:\\Converters не найдено .xlsx файлов для конвертации.")
else:
    print(f"\n✨ Все файлы конвертированы и сохранены в папке: '{FOLDER_PATH}'")

input("\nНажмите Enter для выхода...")